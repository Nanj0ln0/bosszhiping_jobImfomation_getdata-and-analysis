import collections
import wordcloud
import re
import pandas as pd
import numpy as np
import os
import openpyxl
import matplotlib.pyplot as plt
import csv

plt.rcParams['font.sans-serif'] = ['SimHei']  # 显示中文标签
plt.rcParams['axes.unicode_minus'] = False  # 设置正常显示符号

def job_excel_to_transform_csv(excel_filenames,transform_filename):
    transform_names = ["岗位","地点","薪资","工作经验","学历","公司","技能"]
    transform_data = []
    city_data = []
    for excel_filename in excel_filenames:
        book = openpyxl.load_workbook(excel_filename)
        # sheet = book.active
        sheet_names = book.get_sheet_names()
        for sheet_name in sheet_names:

            if not sheet_name.startswith("Sheet"):

                city_data.append(sheet_name)

                sheet = book.get_sheet_by_name(sheet_name)
                sheet_title = sheet.title
                row = sheet.max_row
                column = sheet.max_column

                for r in range(row):
                    job_names = [
                        "职位名称", "职业标签", "岗位位置", "岗位待遇", "岗位要求", "公司名称", "公司标签", "公司福利", "详情网址", "boss", "boss岗位", "岗位描述"
                    ]
                    job_values = []
                    for c in range(column):
                        val = sheet.cell(row=r+1,column=c+1)
                        val = val.value
                        job_values.append(val)

                    if len(job_names) == len(job_values):
                        val3 = job_values[3]  # 薪资
                        if val3 and val3.find("K") > -1:
                            val4 = job_values[4]
                            if val4 and val4.startswith("[") and val4.endswith("]"):
                                val4 = val4[1:-1]
                                val4s = val4.split(",")
                                if len(val4s) == 2:

                                    transform_item = [
                                        job_values[0],
                                        job_values[2],
                                        job_values[3],
                                        val4s[0],
                                        val4s[1],# 工作经验，学历
                                        job_values[5],
                                        job_values[1]
                                    ]
                                    transform_data.append(transform_item)
                                    # print(transform_item)
        book.close()

    # 接下来讲转换后的数据写入到转换文件中
    df_data = pd.DataFrame(columns=transform_names)
    index = 1
    for transform_item in transform_data:
        df_data.loc[index, :] = transform_item
        index += 1
    df_data.to_csv(transform_filename, index=False, encoding='gb18030')

    return city_data
def get_job_data_from_transform_csv(transform_filename,transform_update_filename):

    data = pd.read_csv(transform_filename, encoding='gb18030')

    data_df = pd.DataFrame(data)
    print("\n查看是否有缺失值\n", data_df.isnull().sum())

    data_df_del_empty = data_df.dropna(subset=['岗位'], axis=0)
    # print("\n删除缺失值‘岗位’的整行\n",data_df_del_empty)
    data_df_del_empty = data_df_del_empty.dropna(subset=['公司'], axis=0)
    # print("\n删除缺失值‘公司’的整行\n",data_df_del_empty)

    print("\n查看是否有缺失值\n", data_df_del_empty.isnull().sum())
    print('去除缺失值后\n', data_df_del_empty)

    # job_data = data_df_del_empty.loc[data_df_del_empty['岗位'].str.contains('Python|python')]
    # print(job_data)#筛选带有python的行
    job_data = data_df_del_empty

    # 区间最小薪资
    job_data_salary = job_data['薪资'].str.split('-', expand=True)[0]
    # print("区间最小薪资（1）：",job_data_salary)  # 区间最小薪资
    # Dataframe新增一列  在第 列新增一列名为' ' 的一列 数据
    job_data.insert(7, '区间最小薪资(K)', job_data_salary)
    # print("区间最小薪资（2）：",job_data)

    # 城市地区
    job_data_location_city = job_data['地点'].str.split('·', expand=True)[0]
    # print("地点（1）",job_data_location_city)  # 北京
    job_data_location_district = job_data['地点'].str.split('·', expand=True)[1]
    # print("地点（2）",job_data_location_district)  # 海淀区

    job_data_location_city_district = []
    for city, district in zip(job_data_location_city, job_data_location_district):
        if not city:
            city = ''
        if not district:
            district = ''

        city_district = city + district
        job_data_location_city_district.append(city_district)

    # print(job_data_location_city_district)  # 北京海淀区
    # Dataframe新增一列  在第 列新增一列名为' ' 的一列 数据
    job_data.insert(8, '城市地区', job_data_location_city_district)
    # print(job_data)
    job_data.insert(9, '城市', job_data_location_city)
    job_data.insert(10, '地区', job_data_location_district)
    job_data.to_csv(transform_update_filename, index=False, encoding='gb18030')


    return job_data
def analysis_job_data(city_data,job_data,out_dir):
    name_prefix = "_".join(city_data)

    def create_dir_not_exist(path):  # 判断文件夹是否存在,不存在-新建
        if not os.path.exists(path):
            os.mkdir(path)
    out_city_dir = out_dir + "/city"

    create_dir_not_exist(out_dir)
    create_dir_not_exist(out_city_dir)

    def draw_bar(row_lable, title):
        figsize_x = 10
        figsize_y = 6
        global list1_education, list2_education, df1, df2
        plt.figure(figsize=(figsize_x, figsize_y))
        list1_education = []
        list2_education = []
        for df1, df2 in job_data.groupby(row_lable):
            list1_education.append(df1)
            list2_education.append(len(df2))
        # print(list1_education)
        # print(list2_education)
        # 利用 * 解包方式 将 一个排序好的元组，通过元组生成器再转成list
        # print(*sorted(zip(list2_education,list1_education)))
        # print(sorted(zip(list2_education,list1_education)))
        # 排序，两个列表对应原始排序,按第几个列表排序，注意先后位置
        list2_education, list1_education = (list(t) for t in zip(*sorted(zip(list2_education, list1_education))))
        plt.bar(list1_education, list2_education)
        plt.title('{}'.format(title))
        plt.savefig('{out_dir}/{title}分析.jpg'.format(out_dir=out_dir,title=title))
        # plt.show()
        plt.close()


    # 学历
    draw_bar('学历', '学历')
    draw_bar('工作经验', '工作经验')
    draw_bar('区间最小薪资(K)', name_prefix+'-薪资分布情况(K)')
    # -----------------------------------------
    # 根据城市地区求均值
    list_group_city1 = []
    list_group_city2 = []

    for df1, df2 in job_data.groupby(job_data['城市地区']):
        # print(df1)
        # print(df2)
        list_group_city1.append(df1)
        salary_list_district = [int(i) for i in (df2['区间最小薪资(K)'].values.tolist())]
        district_salary_mean = round(np.mean(salary_list_district), 2)  # 每个区县的平均薪资 round(a, 2)保留2位小数
        list_group_city2.append(district_salary_mean)
        list_group_city2, list_group_city1 = (list(t) for t in
                                              zip(*sorted(zip(list_group_city2, list_group_city1), reverse=False)))
    #
    # print(list_group_city1)
    # print(list_group_city2)

    plt.figure(figsize=(10, 50))
    plt.barh(list_group_city1, list_group_city2)
    # 坐标轴上的文字说明
    for ax, ay in zip(list_group_city1, list_group_city2):
        # 设置文字说明 第一、二个参数：坐标轴上的值； 第三个参数：说明文字；ha:垂直对齐方式；va：水平对齐方式
        plt.text(ay, ax, '%.2f' % ay, ha='center', va='bottom')
    plt.title(name_prefix+'-各区县招聘工资情况(K)')
    plt.savefig('{out_dir}/{name_prefix}-各区县招聘工资情况(K).jpg'.format(out_dir=out_dir,name_prefix=name_prefix))
    # plt.show()
    plt.close()

    # -----------------------------------------
    # 根据城市分组排序，

    list_group_city11 = []
    list_group_city22 = []
    list_group_city33 = []
    list_group_city44 = []

    for df_city1, df_city2 in job_data.groupby(job_data['城市']):
        # print(df_city1)#市
        # print(df_city2)
        list_group_district2 = []  # 区县列表
        district_mean_salary2 = []  # 工资均值列表
        for df_district1, df_district2 in df_city2.groupby(job_data['地区']):
            # print(df_district1)#区县
            # print(df_district2)#工作
            list_group_district2.append(df_district1)  # 记录区县
            salary_list_district2 = [int(i) for i in (df_district2['区间最小薪资(K)'].values.tolist())]  # 工资列表
            district_salary_mean2 = round(np.mean(salary_list_district2), 2)  # 每个区县的平均薪资 round(a, 2)保留2位小数
            district_mean_salary2.append(district_salary_mean2)  # 记录区县的平均工作的列表

        district_mean_salary2, list_group_district2 = (list(tt) for tt in zip(
            *sorted(zip(district_mean_salary2, list_group_district2), reverse=True)))
        plt.figure(figsize=(10, 6))
        plt.bar(list_group_district2, district_mean_salary2)

        # 坐标轴上的文字说明
        for ax, ay in zip(list_group_district2, district_mean_salary2):
            # 设置文字说明 第一、二个参数：坐标轴上的值； 第三个参数：说明文字；ha:垂直对齐方式；va：水平对齐方式
            plt.text(ax, ay, '%.2f' % ay, ha='center', va='bottom')

        plt.title('{}-各区县招聘工资情况_{}(K)'.format(name_prefix,df_city1))
        plt.savefig('{out_city_dir}/{df_city1}_各区县招聘工资情况(K).jpg'.format(out_city_dir=out_city_dir,df_city1=df_city1))
        # plt.show()
        plt.close()

    # ----------------------------------------------------


    skill_all = job_data['技能']
    print(skill_all)

    skill_list = []

    for i in skill_all:
        # print(type(i))
        print(i)
        # print(i.split(", | ' | \[ | \]  |  \" | "))
        result = re.split(r'[,\' \[, \]  ]', i)
        print(result)
        # if type(i) == list:
        skill_list = skill_list + result
    print('++++++++++++++++++++++++++++++++')
    # print(skill_list)

    list_new = skill_list

    # 词频统计
    word_counts = collections.Counter(list_new)  # 对分词做词频统计
    word_counts_top10 = word_counts.most_common(30)  # 获取前10最高频的词
    # print (word_counts_top10) # 输出检查
    # print (word_counts_top10[0][0]) # 输出检查

    # 生成柱状图
    list_x = []
    list_y = []
    for i in word_counts_top10:
        list_x.append(i[0])
        list_y.append(i[1])
    print('list_x', list_x[1:])
    print('list_y', list_y[1:])
    plt.figure(figsize=(30, 5))
    plt.bar(list_x[1:], list_y[1:])
    plt.savefig(out_dir+'/技能栈_词频_柱状图.png')
    # plt.show()
    plt.close()

    list_new = " ".join(list_new)  # 列表转字符串，以空格间隔
    # print(list_new)


    wc = wordcloud.WordCloud(
        width=800,
        height=600,
        background_color="#ffffff",  # 设置背景颜色
        max_words=50,  # 词的最大数（默认为200）
        max_font_size=60,  # 最大字体尺寸
        min_font_size=10,  # 最小字体尺寸（默认为4）
        # colormap='bone',  # string or matplotlib colormap, default="viridis"
        colormap='hsv',  # string or matplotlib colormap, default="viridis"
        random_state=20,  # 设置有多少种随机生成状态，即有多少种配色方案
        # mask=plt.imread("mask2.gif"),  # 读取遮罩图片！！
        font_path='simhei.ttf'
    )
    my_wordcloud = wc.generate(list_new)

    plt.imshow(my_wordcloud)
    plt.axis("off")
    # plt.show()
    wc.to_file(out_dir+'/技能栈_词云.png')  # 保存图片文件
    plt.close()

if __name__ == '__main__':

    excel_filenames = [
        'data/c++_20221230_北上广深.xlsx',
        'data/c++_20230101_南京青岛.xlsx'
    ]
    out_dir = "data/analysis_c++"
    transform_filename = "data/transform.csv"
    transform_update_filename = "data/transform_update.csv"
    
    city_data = job_excel_to_transform_csv(excel_filenames,transform_filename)
    print("city_data:",city_data)
    job_data = get_job_data_from_transform_csv(transform_filename,transform_update_filename)

    analysis_job_data(city_data,job_data,out_dir)
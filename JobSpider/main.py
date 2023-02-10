import random
import time
import os
from bs4 import BeautifulSoup
from datetime import datetime
from utils.base import base
from utils.MyLogger import MyLoggerInit
import openpyxl
from openpyxl.styles import Font

logger = MyLoggerInit(log_dir="logs")
root_path = os.path.dirname(os.path.abspath(__file__))
chrome_driver_path = root_path + "/bin/chromedriver_win32_242.exe"

class ZhiPin(base):
    def __init__(self):
        super(ZhiPin, self).__init__(logger=logger, proxy_radio=0, chrome_driver_path=chrome_driver_path)
    def __parse_page_job_ul_li(self,soup_job_ul_li):
        parse_state = True

        job_url = None # https://www.zhipin.com/job_detail/xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
        job_name = None # 音视频开发
        job_area = None # 北京-朝阳区

        job_info_salary = None # 30-35K
        job_info_tags = [] # [3-5年,本科]

        try:
            body = soup_job_ul_li.find("div",class_="job-card-body")

            body_a_left = body.find("a",class_="job-card-left")
            job_url = "https://www.zhipin.com"+body_a_left.get("href")

            body_a_left_job_title = body_a_left.find("div",class_="job-title")
            job_name = body_a_left_job_title.find("span",class_="job-name").get_text().strip()
            job_area = body_a_left_job_title.find("span",class_="job-area").get_text().strip()

            body_a_left_job_info = body_a_left.find("div", class_="job-info")
            job_info_salary = body_a_left_job_info.find("span",class_="salary").get_text().strip()
            tags = body_a_left_job_info.find("ul",class_="tag-list").find_all("li")
            for tag in tags:
                job_info_tags.append(tag.get_text().strip())


        except Exception as e:
            self.logger.error("__parse_job_ul_li error:%s", str(e))
            parse_state = False

        company_name = None # 公司名称
        company_tags = [] # 公司标签 ['社交网络', 'D轮及以上', '10000人以上']

        try:
            body = soup_job_ul_li.find("div", class_="job-card-body")

            body_div_right = body.find("div", class_="job-card-right")
            body_div_right_company_right = body_div_right.find("div",class_="company-info")
            company_name = body_div_right_company_right.find("h3",class_="company-name").get_text().strip()

            company_ul = body_div_right_company_right.find("ul",class_="company-tag-list")
            company_ul_lis = company_ul.find_all("li")
            for company_ul_li in company_ul_lis:
                company_tag = company_ul_li.get_text().strip()
                company_tags.append(company_tag)


        except Exception as e:
            self.logger.error("__parse_job_ul_li error:%s", str(e))
            parse_state = False

        job_tags = [] # 职业标签 ['C++', 'Go', '微服务架构', '音视频编解码', '流媒体技术', '音视频开发']
        company_benefits = None # 公司统一福利描述
        try:
            footer = soup_job_ul_li.find("div",class_="job-card-footer")
            footer_tag_list_ul = footer.find("ul",class_="tag-list")
            footer_tag_list_ul_lis = footer_tag_list_ul.find_all("li")
            for footer_tag_list_ul_li in footer_tag_list_ul_lis:
                job_tag = footer_tag_list_ul_li.get_text().strip()
                job_tags.append(job_tag)
            company_benefits = footer.find("div",class_="info-desc").get_text().strip()
        except Exception as e:
            self.logger.error("__parse_job_ul_li error:%s",str(e))
            parse_state = False

        job = {
            "job_name":job_name,
            "job_tags": job_tags,
            "job_area":job_area,
            "job_info_salary":job_info_salary,
            "job_info_tags":job_info_tags,
            "company_name":company_name,
            "company_tags":company_tags,
            "company_benefits":company_benefits,
            "job_url": job_url
        }
        return parse_state,job

    def __get_job_detail(self,url):
        self.logger.info("%s",url)
        parse_state = True
        job_desc = None # 岗位描述：xxx
        job_boss_name = None # 刘女士\n刚刚活跃
        job_boss_attr = None # 海南钦诚·招聘专员

        try:
            self.driver.get(url)
            time.sleep(random.randint(5, 20))
            source = self.driver.page_source
            source_len = len(source)

            soup = BeautifulSoup(source, "lxml")

            soup_job_detail = soup.find("div",class_="job-detail")
            soup_job_detail_section = soup_job_detail.find("div",class_="job-detail-section")

            job_desc = soup_job_detail_section.find("div",class_="job-sec-text").get_text().strip()
            soup_job_detail_section_boss = soup_job_detail_section.find("div",class_="job-boss-info")
            job_boss_name = soup_job_detail_section_boss.find("h2",class_="name").get_text().strip()
            job_boss_attr = soup_job_detail_section_boss.find("div",class_="boss-info-attr").get_text().strip()
        except Exception as e:
            self.logger.error("__get_job_detail error:%s", str(e))
            parse_state = False

        detail = {
            "job_boss_name":job_boss_name,
            "job_boss_attr":job_boss_attr,
            "job_desc": job_desc
        }

        return parse_state,detail
    def __get_job_page(self, url):
        self.logger.info("%s",url)
        jobs = []
        try:
            self.driver.get(url)
            time.sleep(random.randint(5,20))
            source = self.driver.page_source
            source_len = len(source)

            soup = BeautifulSoup(source, "lxml")
            soup_job_ul = soup.find("ul",class_="job-list-box")
            if soup_job_ul:
                soup_job_ul_lis = soup_job_ul.find_all("li",class_="job-card-wrapper")
                for soup_job_ul_li in soup_job_ul_lis:
                    parse_state,job = self.__parse_page_job_ul_li(soup_job_ul_li)
                    if parse_state:
                        jobs.append(job)

            for job in jobs:
                job_url = job.get("job_url")
                parse_state,detail = self.__get_job_detail(url=job_url)
                if parse_state:
                     print(detail)
                     job['job_boss_name'] = detail.get("job_boss_name")
                     job['job_boss_attr'] = detail.get("job_boss_attr")
                     job['job_desc'] = detail.get("job_desc")
        except Exception as e:
            self.logger.error("__get_job_page error:%s", str(e))
        return jobs

    def search(self,keyword,citys,pages):

        # time.sleep(1)
        # self.driver.get("https://www.baidu.com/")
        time.sleep(0.5)
        self.driver.get("https://www.zhipin.com/")
        time.sleep(0.5)

        # parse_state,detail = self.__get_job_detail(url="https://www.zhipin.com/job_detail/e089f0cefa95815a1XJ62tm4GVdY.html?lid=1sjvvBuKEFN.search.1&securityId=gMT45Iqq5mzmV-z14Pezxjg7jPmlPE1hDDXi2HKfE-wWovMXrdDDxLJC_HJIBi3qtT2SpHcg-09A4_s5_QPsUd0EkOQh2nld-RzPrT6LFITUVb_rjxCe0IdZFv7OUTVlgoFLTcpSgdWA&sessionId=")
        # print(parse_state,detail)
        # exit(0)

        book = openpyxl.Workbook()
        # book_date = datetime.now().strftime("%Y%m%d_%H%M%S")
        book_date = datetime.now().strftime("%Y%m%d")
        citys_name = []
        for index,city in enumerate(citys):
            city_name, city_code = city
            citys_name.append(city_name)

            sh = book.create_sheet(city_name, index)

            self.driver.get(f"https://www.zhipin.com/web/geek/job?city={city_code}".format(city_code=city_code))
            time.sleep(2)
            job_row_num = 1

            job_names = {
                "job_name": ["职位名称",30],
                "job_tags": ["职业标签",30],
                "job_area": ["岗位位置",15],
                "job_info_salary": ["岗位待遇",15],
                "job_info_tags": ["岗位要求",20],
                "company_name": ["公司名称",30],
                "company_tags": ["公司标签",40],
                "company_benefits": ["公司福利",40],
                "job_url": ["详情网址",10],
                'job_boss_name': ["boss",20],
                'job_boss_attr': ["boss岗位",30],
                'job_desc': ["岗位描述",60],
            }
            chars = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            # 设置excel列 start
            job_column_num = 1
            for job_name_en, attrs in job_names.items():
                sh.cell(job_row_num, job_column_num).value = str(attrs[0])
                sh.column_dimensions[chars[job_column_num-1]].width = attrs[1]
                sh.column_dimensions[chars[job_column_num-1]].font = Font(bold=True)
                job_column_num += 1
            job_row_num += 1
            # 设置excel列 end

            for page in pages:
                url = "https://www.zhipin.com/web/geek/job?query={keyword}&city={city_code}&page={page}".format(
                    keyword=keyword,
                    city_code=city_code,
                    page=page
                )
                jobs = self.__get_job_page(url=url)
                for job in jobs:
                    job_column_num = 1
                    for job_key,job_value in job.items():
                        sh.cell(job_row_num,job_column_num).value = str(job_value)
                        job_column_num += 1
                    job_row_num += 1

                print(len(jobs),jobs)
            time.sleep(30)

        book.save('{keyword}_{book_date}_{citys_name}.xlsx'.format(
            keyword=keyword,
            book_date=book_date,
            citys_name="".join(citys_name)
            )
        )

if __name__ == '__main__':
    # keyword = "C++开发"
    # keyword = "Java开发"
    keyword = "Python"
    keyword = "图像算法"
    # keyword = "旅游"
    # keyword = "测试工程师"
    # keyword = "运维"
    # keyword = "运营"
    # keyword = ".NET"
    # keyword = "大数据开发"
    # keyword = "算法"
    # keyword = "前端开发"
    citys = [
        # ("全国", "100010000"),
        # ("北京", "101010100"),
        # ("上海", "101020100"),
        # ("广州", "101280100"),
        # ("深圳", "101280600"),
        # ("杭州", "101210100"),
        ("南京", "101190100"),
        # ("青岛", "101120200"),
        # ("成都", "101270100"),
        # ("重庆", "101040100"),
        # ("苏州", "101190400"),
        # ("武汉", "101200100"),
        # ("长沙", "101250100"),
        # ("西安", "101110100"),
        # ("济南", "101120100"),
        # ("厦门", "101230200"),

    ]
    pages = [1,2,3,4,5,6,7,8,9,10] # 搜索该城市下，该关键词对应的招聘岗位的数据页码
    # pages = [1] # 搜索该城市下，该关键词对应的招聘岗位的数据页码

    spider = ZhiPin()
    spider.startWebDriver()

    for city in citys:
        spider.search(keyword=keyword,citys=[city],pages=pages)

    spider.stopWebDriver()

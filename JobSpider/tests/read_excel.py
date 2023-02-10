import openpyxl

def read_excel1():
    book = openpyxl.load_workbook('../data/java/java_20230112_北京上海.xlsx')
    # sheet = book.active
    sheet_names = book.get_sheet_names()
    for sheet_name in sheet_names:

        if not sheet_name.startswith("Sheet"):
            sheet = book.get_sheet_by_name(sheet_name)
            sheet_title = sheet.title
            row = sheet.max_row
            column = sheet.max_column

            for r in range(row):
                vals = []
                for c in range(column):
                    val = sheet.cell(row=r+1,column=c+1)
                    val = val.value
                    vals.append(val)
                print(sheet_name,sheet_title,vals)

        # break
        # exit(0)

    book.close()


if __name__ == '__main__':
    read_excel1()


